#!/usr/bin/env python3
"""
Generate a master Excel file with 5 tabs:
  1. General     – profiles from data/profiles/
  2. Enterprise  – profiles from data-enterprise-seed/profiles/
  3. Fintech     – profiles from data-fintech-seed/profiles/
  4. SaaS        – profiles from data-saas/profiles/
  5. All Profiles – every profile from all 4 sources combined
"""

import json
import glob
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE_DIR, "all_investors_master.xlsx")

# Data sources: (sheet_name, profiles_dir, urls_file_or_None)
SOURCES = [
    ("General",    os.path.join(BASE_DIR, "data", "profiles"),                    None),
    ("Enterprise", os.path.join(BASE_DIR, "data-enterprise-seed", "profiles"),    os.path.join(BASE_DIR, "data-enterprise-seed", "all_investor_urls.json")),
    ("Fintech",    os.path.join(BASE_DIR, "data-fintech-seed", "profiles"),       os.path.join(BASE_DIR, "data-fintech-seed", "all_investor_urls.json")),
    ("SaaS",       os.path.join(BASE_DIR, "data-saas", "profiles"),              os.path.join(BASE_DIR, "data-saas", "all_investor_urls.json")),
]

# ---------- Columns for each category sheet (expanded arrays) ----------
BASE_HEADERS = [
    "slug", "matched_url", "profile_url", "profile_picture", "scraped_at",
    "name", "location", "signal_score", "position_and_firm", "website", "investor_types",
    "current_firm", "current_firm_url", "current_position",
    "investment_range", "sweet_spot", "fund_size", "investments_on_record",
    "linkedin", "twitter", "angellist", "crunchbase", "social_website",
]

# ---------- Columns for the "All Profiles" sheet (pipe-separated arrays) ----------
ALL_HEADERS = [
    "source", "slug", "profile_url", "name", "location", "investor_types",
    "signal_score", "current_position", "current_firm", "current_firm_url",
    "position_and_firm", "investment_range", "sweet_spot", "fund_size",
    "investments_on_record", "website", "profile_picture",
    "linkedin", "twitter", "crunchbase", "angellist", "social_website",
    "sector_rankings", "experience", "investments", "scraped_at",
]


def load_url_map(urls_file):
    """Load slug -> url mapping from all_investor_urls.json."""
    if not urls_file or not os.path.exists(urls_file):
        return {}
    with open(urls_file, "r", encoding="utf-8") as f:
        url_list = json.load(f)
    return {item["slug"]: item["url"] for item in url_list}


def parse_profile(filepath):
    """Read and return a profile JSON, or None on error."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"  Skipping {os.path.basename(filepath)}: {e}")
        return None


def extract_base_row(data, url_map):
    """Extract the flat base fields from a profile JSON."""
    slug = data.get("slug", "")
    basic = data.get("basicInfo", {}) or {}
    investing = data.get("investingProfile", {}) or {}
    socials = data.get("socials", {}) or {}

    cp = investing.get("currentPosition")
    if isinstance(cp, dict):
        current_position = cp.get("position", "")
        current_firm = cp.get("firm", "")
        current_firm_url = cp.get("firmUrl", "")
    elif isinstance(cp, str):
        current_position = cp
        current_firm = ""
        current_firm_url = ""
    else:
        current_position = ""
        current_firm = ""
        current_firm_url = ""

    investor_types = basic.get("investorTypes", [])

    return {
        "slug": slug,
        "matched_url": url_map.get(slug, ""),
        "profile_url": data.get("profileUrl", ""),
        "profile_picture": data.get("profilePicture", ""),
        "scraped_at": data.get("scraped_at", ""),
        "name": basic.get("name", ""),
        "location": basic.get("location", ""),
        "signal_score": basic.get("signalScore", ""),
        "position_and_firm": basic.get("positionAndFirm", ""),
        "website": basic.get("website", ""),
        "investor_types": "; ".join(investor_types) if investor_types else "",
        "current_firm": current_firm,
        "current_firm_url": current_firm_url,
        "current_position": current_position,
        "investment_range": investing.get("investmentRange", ""),
        "sweet_spot": investing.get("sweetSpot", ""),
        "fund_size": investing.get("fundSize", ""),
        "investments_on_record": investing.get("investmentsOnRecord", ""),
        "linkedin": socials.get("linkedin", ""),
        "twitter": socials.get("twitter", ""),
        "angellist": socials.get("angellist", ""),
        "crunchbase": socials.get("crunchbase", ""),
        "social_website": socials.get("website", ""),
    }


def extract_all_row(data, source_name):
    """Extract a flat row for the 'All Profiles' sheet (pipe-separated lists)."""
    basic = data.get("basicInfo", {}) or {}
    investing = data.get("investingProfile", {}) or {}
    socials = data.get("socials", {}) or {}

    cp = investing.get("currentPosition")
    if isinstance(cp, dict):
        current_position = cp.get("position", "")
        current_firm = cp.get("firm", "")
        current_firm_url = cp.get("firmUrl", "")
    elif isinstance(cp, str):
        current_position = cp
        current_firm = ""
        current_firm_url = ""
    else:
        current_position = ""
        current_firm = ""
        current_firm_url = ""

    types = basic.get("investorTypes", [])
    rankings = data.get("sectorRankings", [])
    exp_list = data.get("experience", [])
    inv_list = data.get("investments", [])

    sector_str = " | ".join(r.get("name", "") for r in rankings) if rankings else ""

    exp_parts = []
    for e in exp_list:
        pos = e.get("position", e.get("title", ""))
        company = e.get("company", "")
        dates = e.get("dates", "")
        if company:
            exp_parts.append(f"{pos} @ {company} ({dates})" if dates else f"{pos} @ {company}")
        elif pos:
            exp_parts.append(f"{pos} ({dates})" if dates else pos)

    inv_parts = []
    for inv in inv_list:
        company = inv.get("company", "")
        stage = inv.get("stage", "")
        date = inv.get("date", "")
        round_size = inv.get("roundSize", "")
        parts = [company]
        if stage: parts.append(stage)
        if date: parts.append(date)
        if round_size: parts.append(round_size)
        inv_parts.append(" - ".join(parts))

    return {
        "source": source_name,
        "slug": data.get("slug", ""),
        "profile_url": data.get("profileUrl", ""),
        "name": basic.get("name", ""),
        "location": basic.get("location", ""),
        "investor_types": " | ".join(types) if types else "",
        "signal_score": basic.get("signalScore", ""),
        "current_position": current_position,
        "current_firm": current_firm,
        "current_firm_url": current_firm_url,
        "position_and_firm": basic.get("positionAndFirm", ""),
        "investment_range": investing.get("investmentRange", ""),
        "sweet_spot": investing.get("sweetSpot", ""),
        "fund_size": investing.get("fundSize", ""),
        "investments_on_record": investing.get("investmentsOnRecord", ""),
        "website": basic.get("website", ""),
        "profile_picture": data.get("profilePicture", ""),
        "linkedin": socials.get("linkedin", ""),
        "twitter": socials.get("twitter", ""),
        "crunchbase": socials.get("crunchbase", ""),
        "angellist": socials.get("angellist", ""),
        "social_website": socials.get("website", ""),
        "sector_rankings": sector_str,
        "experience": " | ".join(exp_parts) if exp_parts else "",
        "investments": " | ".join(inv_parts) if inv_parts else "",
        "scraped_at": data.get("scraped_at", ""),
    }


def style_header(ws, num_cols):
    """Apply styling to the header row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="000000"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def auto_width(ws, max_width=50):
    """Auto-fit column widths based on content (capped)."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(1, min(ws.max_row + 1, 102)):  # sample first 100 rows
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def build_category_sheet(wb, sheet_name, profiles_dir, urls_file):
    """
    Build one sheet with expanded columns for experience/investments/sectors.
    Returns list of (data_dict, source_name) for the All Profiles sheet.
    """
    if not os.path.isdir(profiles_dir):
        print(f"  [{sheet_name}] Directory not found: {profiles_dir} — skipping")
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["No data found"])
        return []

    url_map = load_url_map(urls_file)
    profile_files = sorted(glob.glob(os.path.join(profiles_dir, "*.json")))
    print(f"  [{sheet_name}] Processing {len(profile_files)} profiles...")

    if not profile_files:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["No profiles found"])
        return []

    # First pass: collect rows and discover max array sizes
    rows = []
    raw_data = []
    max_exp = 0
    max_inv = 0
    max_sec = 0

    for fp in profile_files:
        data = parse_profile(fp)
        if data is None:
            continue
        raw_data.append(data)
        row = extract_base_row(data, url_map)
        exp = data.get("experience", [])
        inv = data.get("investments", [])
        sec = data.get("sectorRankings", [])
        max_exp = max(max_exp, len(exp))
        max_inv = max(max_inv, len(inv))
        max_sec = max(max_sec, len(sec))
        row["_exp"] = exp
        row["_inv"] = inv
        row["_sec"] = sec
        rows.append(row)

    # Build headers
    headers = list(BASE_HEADERS)
    for i in range(1, max_exp + 1):
        headers.extend([f"experience_{i}_company", f"experience_{i}_position", f"experience_{i}_dates"])
    for i in range(1, max_inv + 1):
        headers.extend([f"investment_{i}_company", f"investment_{i}_stage", f"investment_{i}_date",
                        f"investment_{i}_round_size", f"investment_{i}_total_raised", f"investment_{i}_co_investors"])
    for i in range(1, max_sec + 1):
        headers.extend([f"sector_ranking_{i}_name", f"sector_ranking_{i}_url"])

    # Create sheet and write
    ws = wb.create_sheet(title=sheet_name)
    ws.append(headers)

    for row in rows:
        exp = row.pop("_exp")
        inv = row.pop("_inv")
        sec = row.pop("_sec")

        for i, e in enumerate(exp, 1):
            row[f"experience_{i}_company"] = e.get("company", "")
            row[f"experience_{i}_position"] = e.get("position", "")
            row[f"experience_{i}_dates"] = e.get("dates", "")

        for i, v in enumerate(inv, 1):
            row[f"investment_{i}_company"] = v.get("company", "")
            row[f"investment_{i}_stage"] = v.get("stage", "")
            row[f"investment_{i}_date"] = v.get("date", "")
            row[f"investment_{i}_round_size"] = v.get("roundSize", "")
            row[f"investment_{i}_total_raised"] = v.get("totalRaised", "")
            co = v.get("coInvestors", [])
            row[f"investment_{i}_co_investors"] = "; ".join(co) if isinstance(co, list) else str(co)

        for i, s in enumerate(sec, 1):
            row[f"sector_ranking_{i}_name"] = s.get("name", "")
            row[f"sector_ranking_{i}_url"] = s.get("url", "")

        ws.append([row.get(h, "") for h in headers])

    style_header(ws, len(headers))
    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    print(f"  [{sheet_name}] {len(rows)} rows, {len(headers)} columns")
    return raw_data


def build_all_sheet(wb, all_data):
    """Build the combined 'All Profiles' sheet."""
    ws = wb.create_sheet(title="All Profiles")
    ws.append(ALL_HEADERS)

    for source_name, data_list in all_data:
        for data in data_list:
            row = extract_all_row(data, source_name)
            ws.append([row.get(h, "") for h in ALL_HEADERS])

    style_header(ws, len(ALL_HEADERS))
    auto_width(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    total = sum(len(d) for _, d in all_data)
    print(f"  [All Profiles] {total} rows, {len(ALL_HEADERS)} columns")


def main():
    print("=" * 60)
    print("  Master Excel Generator — NFX Signal Investor Profiles")
    print("=" * 60)

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    all_data = []

    for sheet_name, profiles_dir, urls_file in SOURCES:
        raw = build_category_sheet(wb, sheet_name, profiles_dir, urls_file)
        all_data.append((sheet_name, raw))

    print()
    build_all_sheet(wb, all_data)

    print(f"\nSaving to {OUTPUT_FILE} ...")
    wb.save(OUTPUT_FILE)

    total = sum(len(d) for _, d in all_data)
    print(f"\nDone! File saved: {OUTPUT_FILE}")
    print(f"Total profiles across all sheets: {total}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
